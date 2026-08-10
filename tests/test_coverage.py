"""Tests for gene coverage computation logic."""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import pandas as pd

from workflow.src.coverage.core import (
    IN_GENE_FILTER,
    compute_insertion_coverage,
    compute_gene_coverage,
    compute_essentiality_coverage,
    compute_characterisation_status_coverage,
    compute_deletion_viability_coverage,
    compute_essentiality_category_coverage,
    coverage_dicts_from_stats_table,
    build_stats_table,
    load_gene_level,
    resolve_duplicate_annotations,
    write_detailed_gene_excel,
)


def _make_insertion_annotation(n_in_gene=30, n_intergenic=10):
    """Synthetic insertion annotation table with required columns."""
    rows = []
    for i in range(n_in_gene):
        rows.append({"Type": "Coding exon", "Distance_to_stop_codon": 10})
    for i in range(n_intergenic):
        rows.append({"Type": "Intergenic region", "Distance_to_stop_codon": 0})
    # Edge: in-gene but too close to stop codon
    rows.append({"Type": "Coding exon", "Distance_to_stop_codon": 3})
    idx = pd.MultiIndex.from_tuples(
        [(f"I", i * 100, "+", f"g{i}") for i in range(len(rows))],
        names=["Chr", "Coordinate", "Strand", "Gene"],
    )
    return pd.DataFrame(rows, index=idx)


def test_in_gene_filter_constant():
    """Exact filter string is preserved from source notebook (quirk)."""
    assert IN_GENE_FILTER == "Type != 'Intergenic region' and Distance_to_stop_codon > 4"


def test_compute_insertion_coverage_counts():
    """In-gene count = rows passing filter; intergenic = complement."""
    annotation = _make_insertion_annotation(n_in_gene=30, n_intergenic=10)
    # 30 in-gene with Distance_to_stop_codon=10, 1 edge with Distance=3 (fails), 10 intergenic
    result = compute_insertion_coverage(annotation)
    assert result["total"] == 41
    assert result["in_gene"] == 30  # edge case excluded
    assert result["intergenic"] == 11


def test_compute_gene_coverage_counts():
    """covered = DR not NaN; not_covered = DR is NaN."""
    gene_result = pd.DataFrame({
        "Systematic ID": ["g1", "g2", "g3", "g4"],
        "DR": [0.5, None, 0.8, None],
        "essentiality": ["E", "V", "E", "V"],
    })
    result = compute_gene_coverage(gene_result)
    assert result["total"] == 4
    assert result["covered"] == 2
    assert result["not_covered"] == 2


def test_compute_essentiality_coverage_excludes_not_determined():
    """Genes with essentiality 'Not_determined' land in neither essential nor non_essential.

    Real releases (e.g. HD_DIT_HAP) carry a third essentiality value besides
    'E'/'V'. Splitting on `== 'E'` vs `== 'V'` (matching the source notebook
    and _HIST_ROW_QUERIES) means such genes are excluded from both buckets,
    so essential.total + non_essential.total < len(gene_result).
    """
    gene_result = pd.DataFrame({
        "Systematic ID": ["g1", "g2", "g3", "g4", "g5"],
        "DR": [0.5, 0.6, 0.7, None, 0.9],
        "essentiality": ["E", "V", "Not_determined", "Not_determined", "E"],
    })
    result = compute_essentiality_coverage(gene_result)
    assert result["essential"]["total"] == 2
    assert result["non_essential"]["total"] == 1
    # 2 (essential) + 1 (non_essential) + 2 (Not_determined) == 5 total genes
    assert result["essential"]["total"] + result["non_essential"]["total"] < len(gene_result)


def test_compute_essentiality_coverage_essential():
    """Essential (E) gene coverage split is correct."""
    # 3 essential genes (g0-g2) with 2 covered (DR not-NaN), 3 non-essential
    # genes (g3-g5) with 2 covered. NOTE: fixed from the original migration
    # plan's fixture, which had DR=[0.5, None, 0.8, None, 0.7, None] — that
    # data has only 1 non-null DR in the V group (g3-g5), making
    # non_essential["covered"] == 2 mathematically unsatisfiable regardless
    # of implementation (3 non-null DR values total across all 6 genes, but
    # the two assertions below sum to 4). Corrected here to g4's DR staying
    # non-null and g3 also covered, matching the assertions' intended shape.
    gene_result = pd.DataFrame({
        "Systematic ID": [f"g{i}" for i in range(6)],
        "DR": [0.5, None, 0.8, 0.6, 0.7, None],
        "essentiality": ["E", "E", "E", "V", "V", "V"],
    })
    result = compute_essentiality_coverage(gene_result)
    assert result["essential"]["total"] == 3
    assert result["essential"]["covered"] == 2
    assert result["non_essential"]["total"] == 3
    assert result["non_essential"]["covered"] == 2


def test_load_gene_level_renames_legacy_um_lam(tmp_path):
    """Legacy um/lam headers are renamed to DR/DL."""
    legacy_tsv = tmp_path / "fitting_results.tsv"
    pd.DataFrame({
        "Systematic ID": ["g1", "g2"],
        "um": [0.5, 0.6],
        "lam": [1.0, 2.0],
        "essentiality": ["E", "V"],
    }).to_csv(legacy_tsv, sep="\t", index=False)

    result = load_gene_level(legacy_tsv)
    assert "DR" in result.columns
    assert "DL" in result.columns
    assert "um" not in result.columns
    assert "lam" not in result.columns
    assert list(result["DR"]) == [0.5, 0.6]
    assert list(result["DL"]) == [1.0, 2.0]


def test_load_gene_level_is_idempotent_when_dr_dl_already_present(tmp_path):
    """Rename only triggers when DR/DL aren't already present — a no-op on current-schema files."""
    current_tsv = tmp_path / "fitting_results.tsv"
    pd.DataFrame({
        "Systematic ID": ["g1", "g2"],
        "DR": [0.5, 0.6],
        "DL": [1.0, 2.0],
        "essentiality": ["E", "V"],
    }).to_csv(current_tsv, sep="\t", index=False)

    result = load_gene_level(current_tsv)
    assert list(result.columns) == ["Systematic ID", "DR", "DL", "essentiality"]
    assert list(result["DR"]) == [0.5, 0.6]
    assert list(result["DL"]) == [1.0, 2.0]


def test_resolve_duplicate_annotations_keeps_passing_duplicate():
    """When one duplicate passes IN_GENE_FILTER and one doesn't, the passing one is kept."""
    idx = pd.MultiIndex.from_tuples(
        [("I", 100, "+", "TTAA"), ("I", 100, "+", "TTAA")],
        names=["Chr", "Coordinate", "Strand", "Target"],
    )
    annotations = pd.DataFrame(
        {
            "Type": ["Intergenic region", "Coding gene"],
            "Distance_to_stop_codon": [0, 10],
        },
        index=idx,
    )
    result = resolve_duplicate_annotations(annotations)
    assert len(result) == 1
    # The row that passes IN_GENE_FILTER (Coding gene, Distance=10) is kept.
    assert result.iloc[0]["Type"] == "Coding gene"
    assert result.iloc[0]["Distance_to_stop_codon"] == 10


def test_resolve_duplicate_annotations_deterministic_when_neither_passes():
    """When neither duplicate passes IN_GENE_FILTER, the first row (original file order) is kept deterministically."""
    idx = pd.MultiIndex.from_tuples(
        [("I", 200, "+", "TTAA"), ("I", 200, "+", "TTAA")],
        names=["Chr", "Coordinate", "Strand", "Target"],
    )
    annotations = pd.DataFrame(
        {
            "Type": ["Intergenic region", "Intergenic region"],
            "Distance_to_stop_codon": [0, 0],
            "_marker": ["first", "second"],
        },
        index=idx,
    )
    result = resolve_duplicate_annotations(annotations)
    assert len(result) == 1
    # Neither passes IN_GENE_FILTER, so the stable sort preserves original
    # order and the first row ("first") is kept — deterministic, not
    # dependent on pandas' default (unstable) sort algorithm.
    assert result.iloc[0]["_marker"] == "first"


def test_resolve_duplicate_annotations_no_duplicates_is_noop():
    """No duplicate index values -> annotations pass through unchanged."""
    idx = pd.MultiIndex.from_tuples(
        [("I", 100, "+", "TTAA"), ("I", 200, "+", "TTAA")],
        names=["Chr", "Coordinate", "Strand", "Target"],
    )
    annotations = pd.DataFrame(
        {"Type": ["Coding gene", "Intergenic region"], "Distance_to_stop_codon": [10, 0]},
        index=idx,
    )
    result = resolve_duplicate_annotations(annotations)
    pd.testing.assert_frame_equal(result, annotations)


def test_compute_characterisation_status_coverage_splits_by_status():
    """Coverage is computed separately for each characterisation_status value."""
    gene_result = pd.DataFrame({
        "Systematic ID": ["g1", "g2", "g3", "g4", "g5", "g6"],
        "DR": [0.5, None, 0.8, 0.6, None, 0.9],
        "characterisation_status": [
            "biological role published",
            "biological role published",
            "biological role inferred",
            "conserved unknown",
            "conserved unknown",
            "dubious",
        ],
        "essentiality": ["E", "V", "E", "V", "E", "V"],
    })
    result = compute_characterisation_status_coverage(gene_result)

    # 2 genes with "biological role published": 1 covered (g1), 1 not covered (g2)
    assert result["biological role published"]["total"] == 2
    assert result["biological role published"]["covered"] == 1
    assert result["biological role published"]["not_covered"] == 1

    # 1 gene with "biological role inferred": 1 covered (g3)
    assert result["biological role inferred"]["total"] == 1
    assert result["biological role inferred"]["covered"] == 1

    # 2 genes with "conserved unknown": 1 covered (g4), 1 not covered (g5)
    assert result["conserved unknown"]["total"] == 2
    assert result["conserved unknown"]["covered"] == 1

    # 1 gene with "dubious": 1 covered (g6)
    assert result["dubious"]["total"] == 1
    assert result["dubious"]["covered"] == 1


def test_compute_characterisation_status_coverage_handles_missing_column():
    """Returns empty dict when characterisation_status column is missing."""
    gene_result = pd.DataFrame({
        "Systematic ID": ["g1", "g2"],
        "DR": [0.5, 0.6],
        "essentiality": ["E", "V"],
    })
    result = compute_characterisation_status_coverage(gene_result)
    assert result == {}


def test_compute_characterisation_status_coverage_skips_null_status():
    """Genes with null characterisation_status are excluded from all categories."""
    gene_result = pd.DataFrame({
        "Systematic ID": ["g1", "g2", "g3"],
        "DR": [0.5, 0.6, 0.7],
        "characterisation_status": ["biological role published", None, "conserved unknown"],
        "essentiality": ["E", "V", "E"],
    })
    result = compute_characterisation_status_coverage(gene_result)

    # Only 2 categories (g2 with null status is excluded)
    assert len(result) == 2
    assert "biological role published" in result
    assert "conserved unknown" in result
    assert result["biological role published"]["total"] == 1
    assert result["conserved unknown"]["total"] == 1


def test_compute_deletion_viability_coverage_splits_by_viability():
    """Coverage is computed separately for each deletion_viability value."""
    gene_result = pd.DataFrame({
        "Systematic ID": ["g1", "g2", "g3", "g4", "g5", "g6"],
        "DR": [0.5, None, 0.8, 0.6, None, 0.9],
        "deletion_viability": [
            "viable", "viable", "inviable", "depends_on_conditions", "depends_on_conditions", "unknown",
        ],
    })
    result = compute_deletion_viability_coverage(gene_result)

    assert result["viable"]["total"] == 2
    assert result["viable"]["covered"] == 1
    assert result["viable"]["not_covered"] == 1

    assert result["inviable"]["total"] == 1
    assert result["inviable"]["covered"] == 1

    assert result["depends_on_conditions"]["total"] == 2
    assert result["depends_on_conditions"]["covered"] == 1

    assert result["unknown"]["total"] == 1
    assert result["unknown"]["covered"] == 1


def test_compute_essentiality_category_coverage_includes_not_determined():
    """Unlike compute_essentiality_coverage, every essentiality value gets its own row."""
    gene_result = pd.DataFrame({
        "Systematic ID": ["g1", "g2", "g3", "g4", "g5"],
        "DR": [0.5, 0.6, 0.7, None, 0.9],
        "essentiality": ["E", "V", "Not_determined", "Not_determined", "E"],
    })
    result = compute_essentiality_category_coverage(gene_result)

    assert set(result) == {"E", "V", "Not_determined"}
    assert result["E"]["total"] == 2
    assert result["E"]["covered"] == 2
    assert result["V"]["total"] == 1
    assert result["V"]["covered"] == 1
    assert result["Not_determined"]["total"] == 2
    assert result["Not_determined"]["covered"] == 1
    # Every gene lands in exactly one bucket (no exclusion, unlike compute_essentiality_coverage).
    assert sum(v["total"] for v in result.values()) == len(gene_result)


# =============================================================================
# STATS-TABLE READBACK (figures read the same numbers the stats rule wrote)
# =============================================================================
def _make_per_chromosome():
    """Minimal per-chromosome insertion table for build_stats_table."""
    return pd.DataFrame([
        {"Chr": "I", "total": 100, "in_gene": 40, "intergenic": 60},
        {"Chr": "II", "total": 80, "in_gene": 30, "intergenic": 50},
    ])


def test_coverage_dicts_from_stats_table_roundtrips_build_stats_table():
    """coverage_dicts_from_stats_table is the inverse of build_stats_table."""
    insertion_coverage = {"total": 180, "in_gene": 70, "intergenic": 110}
    gene_coverage = {"total": 50, "covered": 40, "not_covered": 10}
    essentiality_coverage = {
        "essential": {"total": 20, "covered": 15, "not_covered": 5},
        "non_essential": {"total": 25, "covered": 20, "not_covered": 5},
    }
    per_chromosome = _make_per_chromosome()
    characterisation_status_coverage = {
        "biological role published": {"total": 30, "covered": 25, "not_covered": 5},
        "conserved unknown": {"total": 12, "covered": 8, "not_covered": 4},
    }
    deletion_viability_coverage = {
        "viable": {"total": 28, "covered": 24, "not_covered": 4},
        "inviable": {"total": 10, "covered": 9, "not_covered": 1},
    }
    essentiality_category_coverage = {
        "E": {"total": 20, "covered": 15, "not_covered": 5},
        "V": {"total": 25, "covered": 20, "not_covered": 5},
        "Not_determined": {"total": 5, "covered": 3, "not_covered": 2},
    }

    stats = build_stats_table(
        insertion_coverage, gene_coverage, essentiality_coverage,
        per_chromosome, characterisation_status_coverage,
        deletion_viability_coverage, essentiality_category_coverage,
    )
    ins, gene, ess, per_chr, char, viability, ess_cat = coverage_dicts_from_stats_table(stats)

    assert ins == insertion_coverage
    assert gene == gene_coverage
    assert ess == essentiality_coverage
    assert char == characterisation_status_coverage
    assert viability == deletion_viability_coverage
    assert ess_cat == essentiality_category_coverage
    # Per-chromosome: Chr labels recovered without the chr_ prefix build_stats_table added.
    assert list(per_chr["Chr"]) == ["I", "II"]
    assert list(per_chr["in_gene"]) == [40, 30]
    assert list(per_chr["intergenic"]) == [60, 50]


def test_coverage_dicts_from_stats_table_missing_row_raises():
    """A stats table missing a required summary row raises a clear error."""
    import pytest

    # Only a gene|all row — insertion|all is absent.
    stats = pd.DataFrame([
        {"metric": "gene", "category": "all", "total": 10, "covered": 8, "not_covered": 2},
    ])
    with pytest.raises(ValueError, match="missing required row"):
        coverage_dicts_from_stats_table(stats)


def test_coverage_dicts_from_stats_table_no_characterisation_rows():
    """Stats table without characterisation_/deletion_viability_/essentiality_ rows yields empty category dicts."""
    insertion_coverage = {"total": 10, "in_gene": 4, "intergenic": 6}
    gene_coverage = {"total": 5, "covered": 3, "not_covered": 2}
    essentiality_coverage = {
        "essential": {"total": 2, "covered": 1, "not_covered": 1},
        "non_essential": {"total": 3, "covered": 2, "not_covered": 1},
    }
    stats = build_stats_table(
        insertion_coverage, gene_coverage, essentiality_coverage, _make_per_chromosome()
    )
    _ins, _gene, _ess, _per_chr, char, viability, ess_cat = coverage_dicts_from_stats_table(stats)
    assert char == {}
    assert viability == {}
    assert ess_cat == {}


def test_build_stats_table_percent_columns():
    """covered_pct / not_covered_pct are covered/not_covered over total, rounded to 1 decimal."""
    insertion_coverage = {"total": 200, "in_gene": 70, "intergenic": 130}
    gene_coverage = {"total": 50, "covered": 40, "not_covered": 10}
    essentiality_coverage = {
        "essential": {"total": 3, "covered": 1, "not_covered": 2},
        "non_essential": {"total": 25, "covered": 20, "not_covered": 5},
    }
    stats = build_stats_table(
        insertion_coverage, gene_coverage, essentiality_coverage, _make_per_chromosome()
    )
    assert "covered_pct" in stats.columns
    assert "not_covered_pct" in stats.columns

    gene_all = stats[(stats["metric"] == "gene") & (stats["category"] == "all")].iloc[0]
    assert gene_all["covered_pct"] == 80.0
    assert gene_all["not_covered_pct"] == 20.0

    ins_all = stats[(stats["metric"] == "insertion") & (stats["category"] == "all")].iloc[0]
    assert ins_all["covered_pct"] == 35.0
    assert ins_all["not_covered_pct"] == 65.0

    # 1/3 -> 33.3, 2/3 -> 66.7 (rounding check)
    essential = stats[(stats["metric"] == "gene") & (stats["category"] == "essential")].iloc[0]
    assert essential["covered_pct"] == 33.3
    assert essential["not_covered_pct"] == 66.7


def test_build_stats_table_percent_handles_zero_total():
    """A category with total == 0 yields NaN percents, not a divide-by-zero error."""
    import numpy as np

    insertion_coverage = {"total": 10, "in_gene": 4, "intergenic": 6}
    gene_coverage = {"total": 5, "covered": 3, "not_covered": 2}
    essentiality_coverage = {
        "essential": {"total": 0, "covered": 0, "not_covered": 0},
        "non_essential": {"total": 3, "covered": 2, "not_covered": 1},
    }
    stats = build_stats_table(
        insertion_coverage, gene_coverage, essentiality_coverage, _make_per_chromosome()
    )
    essential = stats[(stats["metric"] == "gene") & (stats["category"] == "essential")].iloc[0]
    assert np.isnan(essential["covered_pct"])
    assert np.isnan(essential["not_covered_pct"])


# =============================================================================
# DETAILED GENE EXCEL (one sheet per characterisation_status / essentiality / deletion_viability value)
# =============================================================================
def _make_detailed_table():
    return pd.DataFrame({
        "Systematic ID": ["g1", "g2", "g3", "g4", "g5", "g6"],
        "Name": ["g1", "g2", "g3", "g4", "g5", "g6"],
        "characterisation_status": [
            "biological role published",
            "biological role published",
            "biological role inferred",
            "conserved unknown",
            "conserved unknown",
            "dubious",
        ],
        "deletion_viability": ["viable", "inviable", "viable", "unknown", "depends_on_conditions", "unknown"],
        "DR": [0.5, None, 0.8, 0.6, None, 0.9],
        "DL": [1.0, None, 2.0, 1.5, None, 3.0],
        "essentiality": ["E", "V", "E", "Not_determined", "V", "Not_determined"],
        "coverage_status": ["covered", "not_covered", "covered", "covered", "not_covered", "covered"],
    })


def test_write_detailed_gene_excel_sheets_cover_every_category(tmp_path):
    """Every characterisation_status, essentiality, and deletion_viability value gets its own sheet."""
    import openpyxl

    detailed_table = _make_detailed_table()
    out_path = tmp_path / "detailed_genes.xlsx"
    write_detailed_gene_excel(detailed_table, out_path)

    wb = openpyxl.load_workbook(out_path, read_only=True)
    assert set(wb.sheetnames) == {
        "All genes",
        "biological role published",
        "biological role inferred",
        "conserved unknown",
        "dubious",
        "essential",
        "non_essential",
        "essentiality_not_determined",
        "viable",
        "inviable",
        "unknown",
        "depends_on_conditions",
    }
    assert len(wb.sheetnames) == 12


def test_write_detailed_gene_excel_sheet_row_counts_match_value_counts(tmp_path):
    """Each category sheet contains exactly the genes with that category value."""
    import openpyxl

    detailed_table = _make_detailed_table()
    out_path = tmp_path / "detailed_genes.xlsx"
    write_detailed_gene_excel(detailed_table, out_path)

    wb = openpyxl.load_workbook(out_path, read_only=True)

    def _row_count(sheet_name):
        return wb[sheet_name].max_row - 1  # exclude header

    assert _row_count("All genes") == 6
    assert _row_count("biological role published") == 2
    assert _row_count("essential") == 2
    assert _row_count("non_essential") == 2
    assert _row_count("essentiality_not_determined") == 2
    assert _row_count("viable") == 2
    assert _row_count("inviable") == 1
    assert _row_count("unknown") == 2
    assert _row_count("depends_on_conditions") == 1


def test_write_detailed_gene_excel_skips_missing_columns(tmp_path):
    """When essentiality/deletion_viability columns are absent, only the characterisation_status sheets are written."""
    import openpyxl

    detailed_table = _make_detailed_table().drop(columns=["essentiality", "deletion_viability"])
    out_path = tmp_path / "detailed_genes.xlsx"
    write_detailed_gene_excel(detailed_table, out_path)

    wb = openpyxl.load_workbook(out_path, read_only=True)
    assert set(wb.sheetnames) == {
        "All genes",
        "biological role published",
        "biological role inferred",
        "conserved unknown",
        "dubious",
    }
    assert len(wb.sheetnames) == 5
